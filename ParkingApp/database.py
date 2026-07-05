"""
مدیریت پایگاه داده SQLite
نسخه 3.0 - Production Ready
"""

import sqlite3
import os
import shutil
from datetime import datetime, timedelta
from contextlib import contextmanager
import threading


class ParkingDatabase:
    """مدیریت پایگاه داده SQLite با پشتیبانی همزمانی"""

    _instance = None
    _lock = threading.Lock()

    def __new__(cls, db_path=None):
        if cls._instance is None:
            with cls._lock:
                if cls._instance is None:
                    cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self, db_path="data/parking.db"):
        if not hasattr(self, 'initialized'):
            self.db_path = db_path
            self._ensure_data_dir()
            self.init_database()
            self.initialized = True

    def _ensure_data_dir(self):
        """اطمینان از وجود پوشه‌های مورد نیاز"""
        directories = ['data', 'backups', 'exports', 'captured_plates', 'logs']
        for directory in directories:
            os.makedirs(directory, exist_ok=True)

    @contextmanager
    def get_connection(self):
        """مدیریت اتصال به پایگاه داده با context manager"""
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA foreign_keys=ON")
        try:
            yield conn
            conn.commit()
        except Exception as e:
            conn.rollback()
            raise e
        finally:
            conn.close()

    def init_database(self):
        """ایجاد جداول پایگاه داده"""
        with self.get_connection() as conn:
            c = conn.cursor()

            # جدول تنظیمات
            c.execute('''CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL,
                description TEXT,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')

            # جدول خودروهای فعال
            c.execute('''CREATE TABLE IF NOT EXISTS active_cars (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plate_number TEXT UNIQUE NOT NULL,
                plate_part1 TEXT NOT NULL,
                plate_letter TEXT NOT NULL,
                plate_part2 TEXT NOT NULL,
                plate_part3 TEXT,
                entry_time TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                entry_image TEXT,
                plate_type TEXT DEFAULT 'personal',
                province TEXT,
                operator_name TEXT,
                notes TEXT
            )''')

            # جدول تاریخچه
            c.execute('''CREATE TABLE IF NOT EXISTS parking_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                plate_number TEXT NOT NULL,
                plate_part1 TEXT,
                plate_letter TEXT,
                plate_part2 TEXT,
                plate_part3 TEXT,
                entry_time TIMESTAMP NOT NULL,
                exit_time TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP,
                duration_minutes INTEGER NOT NULL,
                duration_hours REAL NOT NULL,
                cost REAL NOT NULL,
                entry_image TEXT,
                exit_image TEXT,
                plate_type TEXT,
                province TEXT,
                discount_type TEXT DEFAULT 'none',
                discount_percent REAL DEFAULT 0,
                final_cost REAL NOT NULL,
                operator_name TEXT,
                payment_method TEXT DEFAULT 'cash',
                payment_status TEXT DEFAULT 'paid',
                notes TEXT
            )''')

            # جدول اپراتورها
            c.execute('''CREATE TABLE IF NOT EXISTS operators (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                full_name TEXT NOT NULL,
                role TEXT DEFAULT 'operator',
                is_active BOOLEAN DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')

            # جدول لاگ رویدادها
            c.execute('''CREATE TABLE IF NOT EXISTS event_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                event_type TEXT NOT NULL,
                plate_number TEXT,
                description TEXT,
                operator_name TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )''')

            # جدول نرخ‌های ویژه
            c.execute('''CREATE TABLE IF NOT EXISTS special_rates (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                plate_type TEXT DEFAULT 'personal',
                discount_percent REAL DEFAULT 0,
                is_active BOOLEAN DEFAULT 1,
                start_time TEXT,
                end_time TEXT,
                description TEXT
            )''')

            # ایجاد ایندکس‌ها
            c.execute('''CREATE INDEX IF NOT EXISTS idx_history_plate 
                        ON parking_history(plate_number)''')
            c.execute('''CREATE INDEX IF NOT EXISTS idx_history_exit_time 
                        ON parking_history(exit_time)''')
            c.execute('''CREATE INDEX IF NOT EXISTS idx_history_date 
                        ON parking_history(date(exit_time))''')
            c.execute('''CREATE INDEX IF NOT EXISTS idx_active_plate 
                        ON active_cars(plate_number)''')
            c.execute('''CREATE INDEX IF NOT EXISTS idx_active_entry 
                        ON active_cars(entry_time)''')

            # درج داده‌های پیش‌فرض
            self._insert_default_settings(c)
            self._insert_default_operator(c)
            self._insert_default_rates(c)

    def _insert_default_settings(self, cursor):
        """درج تنظیمات پیش‌فرض"""
        defaults = [
            ('parking_name', 'پارکینگ اصلی', 'نام پارکینگ'),
            ('hourly_rate', '5000', 'نرخ هر ساعت (تومان)'),
            ('free_minutes', '15', 'دقایق رایگان'),
            ('max_daily_cost', '50000', 'سقف هزینه روزانه (تومان)'),
            ('currency_unit', 'تومان', 'واحد پول'),
            ('backup_enabled', 'true', 'فعال بودن بکاپ خودکار'),
            ('backup_interval_hours', '24', 'فاصله بکاپ (ساعت)'),
            ('company_name', 'شرکت مدیریت پارکینگ', 'نام شرکت'),
            ('company_phone', '', 'تلفن شرکت'),
            ('report_header', 'صورتحساب پارکینگ', 'عنوان گزارش'),
            ('camera_index', '0', 'شاخص دوربین'),
            ('camera_resolution', '640x480', 'رزولوشن دوربین'),
            ('auto_detect_plate', 'true', 'تشخیص خودکار پلاک'),
            ('theme', 'طلایی-سرمه‌ای', 'تم برنامه'),
            ('font_size', '9', 'اندازه فونت'),
        ]

        for key, value, desc in defaults:
            cursor.execute('''INSERT OR IGNORE INTO settings (key, value, description) 
                            VALUES (?, ?, ?)''', (key, value, desc))

    def _insert_default_operator(self, cursor):
        """درج اپراتور پیش‌فرض"""
        cursor.execute('''INSERT OR IGNORE INTO operators (username, full_name, role) 
                        VALUES (?, ?, ?)''', ('admin', 'مدیر سیستم', 'admin'))

    def _insert_default_rates(self, cursor):
        """درج نرخ‌های ویژه پیش‌فرض"""
        # حذف نرخ‌های قدیمی
        cursor.execute("DELETE FROM special_rates")

        # نرخ‌های پیش‌فرض
        rates = [
            {
                'name': 'تاکسی',
                'plate_type': 'taxi',
                'discount_percent': 30,
                'is_active': 1,
                'start_time': None,
                'end_time': None,
                'description': 'تخفیف ویژه تاکسی‌ها'
            },
            {
                'name': 'دولتی',
                'plate_type': 'governmental',
                'discount_percent': 30,
                'is_active': 1,
                'start_time': None,
                'end_time': None,
                'description': 'تخفیف خودروهای دولتی'
            },
            {
                'name': 'شبانه',
                'plate_type': 'personal',
                'discount_percent': 20,
                'is_active': 1,
                'start_time': '22:00',
                'end_time': '06:00',
                'description': 'تخفیف شبانه (۲۲ تا ۶ صبح)'
            }
        ]

        for rate in rates:
            try:
                cursor.execute('''
                    INSERT INTO special_rates 
                    (name, plate_type, discount_percent, is_active, 
                     start_time, end_time, description)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (
                    rate['name'],
                    rate['plate_type'],
                    rate['discount_percent'],
                    rate['is_active'],
                    rate['start_time'],
                    rate['end_time'],
                    rate['description']
                ))
            except Exception as e:
                print(f"Warning: Could not insert rate '{rate['name']}': {e}")

    # ==================== عملیات اصلی ====================

    def car_entry(self, plate_data):
        """ثبت ورود خودرو"""
        with self.get_connection() as conn:
            c = conn.cursor()

            # بررسی تکراری نبودن
            c.execute("SELECT id FROM active_cars WHERE plate_number = ?",
                      (plate_data['plate_number'],))
            if c.fetchone():
                raise ValueError(f"خودرو با پلاک {plate_data['plate_number']} قبلاً ثبت شده است")

            # ثبت ورود
            c.execute('''INSERT INTO active_cars (
                plate_number, plate_part1, plate_letter, plate_part2, plate_part3,
                entry_time, entry_image, plate_type, province, operator_name, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                plate_data['plate_number'],
                plate_data['plate_part1'],
                plate_data['plate_letter'],
                plate_data['plate_part2'],
                plate_data.get('plate_part3', ''),
                datetime.now().isoformat(),
                plate_data.get('entry_image'),
                plate_data.get('plate_type', 'personal'),
                plate_data.get('province', ''),
                plate_data.get('operator_name', ''),
                plate_data.get('notes', '')
            ))

            # ثبت لاگ
            self._log_event(c, 'entry', plate_data['plate_number'],
                            'ورود خودرو', plate_data.get('operator_name', ''))

            return c.lastrowid

    def car_exit(self, plate_number, exit_data=None):
        """ثبت خروج خودرو و محاسبه هزینه"""
        if exit_data is None:
            exit_data = {}

        with self.get_connection() as conn:
            c = conn.cursor()

            # دریافت اطلاعات ورود
            c.execute("SELECT * FROM active_cars WHERE plate_number = ?", (plate_number,))
            car = c.fetchone()

            if not car:
                raise ValueError(f"خودرو با پلاک {plate_number} یافت نشد")

            # محاسبه مدت توقف
            entry_time = datetime.fromisoformat(car['entry_time'])
            exit_time = datetime.now()
            duration = exit_time - entry_time
            total_minutes = int(duration.total_seconds() / 60)
            total_hours = total_minutes / 60

            # دریافت تنظیمات
            hourly_rate = float(self.get_setting('hourly_rate', '5000'))
            free_minutes = int(self.get_setting('free_minutes', '15'))
            max_daily = float(self.get_setting('max_daily_cost', '50000'))

            # محاسبه هزینه
            if total_minutes <= free_minutes:
                cost = 0
                discount_type = 'free_short_stop'
                discount_percent = 100
            else:
                # محاسبه پایه
                hours_charged = max(1, int(total_hours + 0.99))
                cost = hours_charged * hourly_rate

                # بررسی سقف روزانه
                if cost > max_daily:
                    cost = max_daily

                # بررسی تخفیف‌های ویژه
                discount_percent = self._calculate_discount(car['plate_type'])
                discount_type = 'special' if discount_percent > 0 else 'none'

                if discount_percent > 0:
                    cost = cost * (1 - discount_percent / 100)

            final_cost = cost

            # ثبت در تاریخچه
            c.execute('''INSERT INTO parking_history (
                plate_number, plate_part1, plate_letter, plate_part2, plate_part3,
                entry_time, exit_time, duration_minutes, duration_hours,
                cost, entry_image, exit_image, plate_type, province,
                discount_type, discount_percent, final_cost,
                operator_name, payment_method, payment_status, notes
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
                car['plate_number'],
                car['plate_part1'],
                car['plate_letter'],
                car['plate_part2'],
                car['plate_part3'],
                car['entry_time'],
                exit_time.isoformat(),
                total_minutes,
                round(total_hours, 2),
                cost,
                car['entry_image'],
                exit_data.get('exit_image'),
                car['plate_type'],
                car['province'],
                discount_type,
                discount_percent,
                final_cost,
                exit_data.get('operator_name', ''),
                exit_data.get('payment_method', 'cash'),
                'paid',
                exit_data.get('notes', '')
            ))

            history_id = c.lastrowid

            # حذف از خودروهای فعال
            c.execute("DELETE FROM active_cars WHERE plate_number = ?", (plate_number,))

            # ثبت لاگ
            self._log_event(c, 'exit', plate_number,
                            f'خروج خودرو - هزینه: {final_cost:,.0f} تومان',
                            exit_data.get('operator_name', ''))

            return {
                'history_id': history_id,
                'plate_number': plate_number,
                'entry_time': car['entry_time'],
                'exit_time': exit_time.isoformat(),
                'duration_minutes': total_minutes,
                'duration_hours': round(total_hours, 2),
                'base_cost': cost,
                'discount_percent': discount_percent,
                'final_cost': final_cost,
                'discount_type': discount_type
            }

    def _calculate_discount(self, plate_type):
        """محاسبه درصد تخفیف بر اساس نوع پلاک و زمان"""
        with self.get_connection() as conn:
            c = conn.cursor()
            current_time = datetime.now().strftime('%H:%M')

            c.execute('''SELECT discount_percent FROM special_rates 
                         WHERE plate_type = ? AND is_active = 1 
                         AND (start_time IS NULL OR start_time <= ?)
                         AND (end_time IS NULL OR end_time >= ?)
                         ORDER BY discount_percent DESC
                         LIMIT 1''',
                      (plate_type, current_time, current_time))

            result = c.fetchone()
            return result['discount_percent'] if result else 0

    # ==================== عملیات خواندن ====================

    def get_active_cars(self):
        """دریافت لیست خودروهای فعال"""
        with self.get_connection() as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM active_cars ORDER BY entry_time DESC")
            cars = []

            for row in c.fetchall():
                car = dict(row)

                # محاسبه مدت حضور
                entry_time = datetime.fromisoformat(car['entry_time'])
                duration = datetime.now() - entry_time
                car['duration_hours'] = round(duration.total_seconds() / 3600, 1)
                car['duration_minutes'] = int(duration.total_seconds() / 60)

                # محاسبه هزینه تقریبی
                hourly_rate = float(self.get_setting('hourly_rate', '5000'))
                free_minutes = int(self.get_setting('free_minutes', '15'))

                if car['duration_minutes'] > free_minutes:
                    hours = max(1, int(car['duration_hours'] + 0.99))
                    car['estimated_cost'] = hours * hourly_rate

                    # اعمال تخفیف
                    discount = self._calculate_discount(car.get('plate_type', 'personal'))
                    if discount > 0:
                        car['estimated_cost'] *= (1 - discount / 100)
                else:
                    car['estimated_cost'] = 0

                cars.append(car)

            return cars

    def get_history(self, page=1, per_page=50, date_from=None, date_to=None,
                    plate_number=None, plate_type=None):
        """دریافت تاریخچه با فیلتر و صفحه‌بندی"""
        with self.get_connection() as conn:
            c = conn.cursor()

            # ساخت query پایه
            query = "FROM parking_history WHERE 1=1"
            params = []

            if date_from:
                query += " AND date(exit_time) >= date(?)"
                params.append(date_from)

            if date_to:
                query += " AND date(exit_time) <= date(?)"
                params.append(date_to)

            if plate_number:
                query += " AND plate_number LIKE ?"
                params.append(f'%{plate_number}%')

            if plate_type:
                query += " AND plate_type = ?"
                params.append(plate_type)

            # تعداد کل
            c.execute(f"SELECT COUNT(*) {query}", params)
            total = c.fetchone()[0]

            # صفحه‌بندی
            offset = (page - 1) * per_page
            c.execute(f"SELECT * {query} ORDER BY exit_time DESC LIMIT ? OFFSET ?",
                      params + [per_page, offset])

            records = [dict(row) for row in c.fetchall()]

            return {
                'records': records,
                'total': total,
                'page': page,
                'per_page': per_page,
                'total_pages': max(1, (total + per_page - 1) // per_page)
            }

    def get_statistics(self, date=None):
        """دریافت آمار"""
        if not date:
            date = datetime.now().strftime('%Y-%m-%d')

        with self.get_connection() as conn:
            c = conn.cursor()

            # آمار روزانه
            c.execute('''SELECT 
                        COUNT(*) as count, 
                        COALESCE(SUM(final_cost), 0) as income,
                        COALESCE(AVG(duration_hours), 0) as avg_duration,
                        COUNT(DISTINCT plate_number) as unique_cars
                        FROM parking_history 
                        WHERE date(exit_time) = ?''', (date,))
            daily = dict(c.fetchone())

            # تفکیک نوع پلاک
            c.execute('''SELECT plate_type, COUNT(*) as count, 
                        COALESCE(SUM(final_cost), 0) as income
                        FROM parking_history 
                        WHERE date(exit_time) = ?
                        GROUP BY plate_type''', (date,))
            by_type = [dict(row) for row in c.fetchall()]

            # ساعات شلوغ
            c.execute('''SELECT strftime('%H', exit_time) as hour, 
                        COUNT(*) as count
                        FROM parking_history 
                        WHERE date(exit_time) = ?
                        GROUP BY hour
                        ORDER BY count DESC
                        LIMIT 5''', (date,))
            peak_hours = [dict(row) for row in c.fetchall()]

            return {
                'daily': daily,
                'by_plate_type': by_type,
                'peak_hours': peak_hours,
                'date': date,
                'active_cars': len(self.get_active_cars())
            }

    def search_plate(self, query):
        """جستجوی پلاک"""
        with self.get_connection() as conn:
            c = conn.cursor()

            # جستجو در خودروهای فعال
            c.execute("SELECT * FROM active_cars WHERE plate_number LIKE ?",
                      (f'%{query}%',))
            active = [dict(row) for row in c.fetchall()]

            # جستجو در تاریخچه
            c.execute('''SELECT * FROM parking_history 
                        WHERE plate_number LIKE ? 
                        ORDER BY exit_time DESC LIMIT 10''',
                      (f'%{query}%',))
            history = [dict(row) for row in c.fetchall()]

            return {'active': active, 'history': history}

    # ==================== تنظیمات ====================

    def get_setting(self, key, default=None):
        """دریافت یک تنظیم"""
        with self.get_connection() as conn:
            c = conn.cursor()
            c.execute("SELECT value FROM settings WHERE key = ?", (key,))
            result = c.fetchone()
            return result['value'] if result else default

    def set_setting(self, key, value):
        """تنظیم یک مقدار"""
        with self.get_connection() as conn:
            c = conn.cursor()
            c.execute('''INSERT OR REPLACE INTO settings (key, value, updated_at) 
                        VALUES (?, ?, CURRENT_TIMESTAMP)''', (key, str(value)))

    def get_all_settings(self):
        """دریافت همه تنظیمات"""
        with self.get_connection() as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM settings")
            return {row['key']: row['value'] for row in c.fetchall()}

    # ==================== لاگ‌ها ====================

    def _log_event(self, cursor, event_type, plate_number, description, operator_name=''):
        """ثبت لاگ رویداد (نیاز به cursor فعال)"""
        cursor.execute('''INSERT INTO event_logs (event_type, plate_number, description, operator_name)
                        VALUES (?, ?, ?, ?)''',
                       (event_type, plate_number, description, operator_name))

    def get_recent_logs(self, limit=100):
        """دریافت لاگ‌های اخیر"""
        with self.get_connection() as conn:
            c = conn.cursor()
            c.execute("SELECT * FROM event_logs ORDER BY created_at DESC LIMIT ?", (limit,))
            return [dict(row) for row in c.fetchall()]

    # ==================== عملیات کمکی ====================

    def backup_database(self):
        """تهیه نسخه پشتیبان"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = f"backups/parking_backup_{timestamp}.db"

        # بستن اتصالات قبل از کپی
        shutil.copy2(self.db_path, backup_path)

        # حذف بکاپ‌های قدیمی (نگه‌داری ۳۰ عدد آخر)
        backups = sorted(os.listdir('backups'))
        if len(backups) > 30:
            for old_backup in backups[:-30]:
                try:
                    os.remove(os.path.join('backups', old_backup))
                except:
                    pass

        return backup_path

    def export_to_excel(self, date_from=None, date_to=None):
        """خروجی Excel از تاریخچه"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "تاریخچه پارکینگ"
            ws.sheet_view.rightToLeft = True

            # هدرها
            headers = [
                'ردیف', 'پلاک', 'نوع', 'استان',
                'زمان ورود', 'زمان خروج', 'مدت (ساعت)',
                'هزینه پایه', 'تخفیف', 'هزینه نهایی',
                'اپراتور', 'روش پرداخت'
            ]

            # استایل هدر
            header_font = Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')

            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # داده‌ها
            history = self.get_history(date_from=date_from, date_to=date_to, per_page=10000)

            for i, record in enumerate(history['records'], 1):
                row_data = [
                    i,
                    record['plate_number'],
                    record.get('plate_type', 'شخصی'),
                    record.get('province', ''),
                    record['entry_time'][:16],
                    record['exit_time'][:16],
                    f"{record['duration_hours']:.1f}",
                    f"{record['cost']:,.0f}",
                    f"{record['discount_percent']}%" if record['discount_percent'] > 0 else "-",
                    f"{record['final_cost']:,.0f}",
                    record.get('operator_name', '-'),
                    record.get('payment_method', 'نقدی')
                ]

                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=i + 1, column=col, value=value)
                    cell.font = Font(name='Tahoma', size=10)
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                    # رنگ‌آمیزی هزینه
                    if col == 10 and record['final_cost'] > 0:
                        cell.font = Font(name='Tahoma', size=10, bold=True, color='c0392b')

            # تنظیم عرض ستون‌ها
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

            # ذخیره
            os.makedirs('exports', exist_ok=True)
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"exports/parking_report_{timestamp}.xlsx"
            wb.save(filename)

            return filename

        except ImportError:
            raise ImportError("کتابخانه openpyxl نصب نیست. لطفاً اجرا کنید: pip install openpyxl")
        except Exception as e:
            raise Exception(f"خطا در خروجی Excel: {str(e)}")

    def get_database_info(self):
        """دریافت اطلاعات پایگاه داده"""
        with self.get_connection() as conn:
            c = conn.cursor()

            info = {}

            # تعداد خودروهای فعال
            c.execute("SELECT COUNT(*) FROM active_cars")
            info['active_cars'] = c.fetchone()[0]

            # تعداد کل تاریخچه
            c.execute("SELECT COUNT(*) FROM parking_history")
            info['total_history'] = c.fetchone()[0]

            # درآمد امروز
            today = datetime.now().strftime('%Y-%m-%d')
            c.execute("SELECT COALESCE(SUM(final_cost), 0) FROM parking_history WHERE date(exit_time) = ?", (today,))
            info['today_income'] = c.fetchone()[0]

            # درآمد ماه
            month = datetime.now().strftime('%Y-%m')
            c.execute("SELECT COALESCE(SUM(final_cost), 0) FROM parking_history WHERE strftime('%Y-%m', exit_time) = ?",
                      (month,))
            info['month_income'] = c.fetchone()[0]

            # حجم پایگاه داده
            if os.path.exists(self.db_path):
                info['db_size'] = os.path.getsize(self.db_path) / (1024 * 1024)  # MB

            # تعداد بکاپ‌ها
            if os.path.exists('backups'):
                info['backup_count'] = len(os.listdir('backups'))

            return info


# ==================== تست ====================

if __name__ == "__main__":
    print("🧪 تست پایگاه داده...")

    try:
        # ایجاد نمونه
        db = ParkingDatabase()
        print("✅ پایگاه داده ایجاد شد")

        # تست ثبت ورود
        try:
            entry_id = db.car_entry({
                'plate_number': '12A345-11',
                'plate_part1': '12',
                'plate_letter': 'الف',
                'plate_part2': '345',
                'plate_part3': '11',
                'plate_type': 'personal',
                'province': 'اصفهان',
                'operator_name': 'admin'
            })
            print(f"✅ ورود ثبت شد: {entry_id}")
        except ValueError as e:
            print(f"⚠️ {e}")

        # تست دریافت خودروهای فعال
        active = db.get_active_cars()
        print(f"✅ خودروهای فعال: {len(active)}")

        # تست ثبت خروج
        try:
            result = db.car_exit('12A345-11', {
                'operator_name': 'admin',
                'payment_method': 'cash'
            })
            print(f"✅ خروج ثبت شد - هزینه: {result['final_cost']:,.0f} تومان")
        except ValueError as e:
            print(f"⚠️ {e}")

        # تست آمار
        stats = db.get_statistics()
        print(f"✅ آمار امروز: {stats['daily']['count']} خودرو")

        # تست تنظیمات
        rate = db.get_setting('hourly_rate')
        print(f"✅ نرخ ساعتی: {rate} تومان")

        # تست بکاپ
        backup_path = db.backup_database()
        print(f"✅ بکاپ در: {backup_path}")

        # تست اطلاعات پایگاه داده
        info = db.get_database_info()
        print(f"✅ اطلاعات: {info}")

        print("\n🎉 همه تست‌ها با موفقیت انجام شد!")

    except Exception as e:
        print(f"❌ خطا: {e}")
        import traceback

        traceback.print_exc()