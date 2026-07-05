"""
تب تاریخچه - نسخه ساده و اسکرول‌دار
"""
from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QFrame, QGroupBox,
    QLineEdit, QDateEdit, QComboBox
)
from PyQt5.QtCore import Qt, pyqtSignal, QDate
from PyQt5.QtGui import QFont, QColor
from datetime import datetime, timedelta
import sys, os

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from plate_utils import IranianPlate


class HistoryTab(QWidget):
    def __init__(self, database):
        super().__init__()
        self.db = database
        self.current_page = 1
        self.per_page = 50
        self.init_ui()
        self.load_data()

    def init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(30, 25, 30, 25)

        # عنوان
        title = QLabel("📊 تاریخچه تردد خودروها")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 26px; font-weight: bold; color: #2c3e50; padding: 15px;")
        layout.addWidget(title)

        # فیلترها
        filter_group = QGroupBox("🔍 فیلتر و جستجو")
        filter_group.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                border: 2px solid #f39c12;
                border-radius: 10px;
                margin-top: 15px;
                padding: 15px;
                padding-top: 35px;
                background-color: white;
            }
            QGroupBox::title {
                left: 15px;
                padding: 8px 20px;
                background-color: #f39c12;
                color: white;
                border-radius: 5px;
            }
        """)

        filter_layout = QHBoxLayout()
        filter_layout.setSpacing(10)

        # جستجوی پلاک
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("🔍 جستجوی پلاک...")
        self.search_input.setStyleSheet("font-size: 14px; padding: 10px; border: 2px solid #ddd; border-radius: 6px;")
        self.search_input.setMinimumHeight(40)

        # تاریخ
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setDate(QDate.currentDate().addDays(-7))
        self.date_from.setStyleSheet("font-size: 14px; padding: 8px; border: 2px solid #ddd; border-radius: 6px;")
        self.date_from.setMinimumHeight(40)

        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setDate(QDate.currentDate())
        self.date_to.setStyleSheet("font-size: 14px; padding: 8px; border: 2px solid #ddd; border-radius: 6px;")
        self.date_to.setMinimumHeight(40)

        # نوع پلاک
        self.type_filter = QComboBox()
        self.type_filter.addItems(["همه", "شخصی", "تاکسی", "دولتی", "نظامی", "سیاسی"])
        self.type_filter.setStyleSheet("font-size: 14px; padding: 8px; border: 2px solid #ddd; border-radius: 6px;")
        self.type_filter.setMinimumHeight(40)

        # دکمه‌ها
        search_btn = QPushButton("🔍 جستجو")
        search_btn.setStyleSheet("""
            QPushButton {
                background-color: #f39c12;
                color: white;
                font-weight: bold;
                padding: 12px 25px;
                border-radius: 6px;
                border: none;
                font-size: 14px;
                min-height: 40px;
            }
            QPushButton:hover { background-color: #e67e22; }
        """)
        search_btn.clicked.connect(self.load_data)

        clear_btn = QPushButton("🔄 پاک کردن")
        clear_btn.setStyleSheet("""
            QPushButton {
                background-color: #95a5a6;
                color: white;
                font-weight: bold;
                padding: 12px 25px;
                border-radius: 6px;
                border: none;
                font-size: 14px;
                min-height: 40px;
            }
            QPushButton:hover { background-color: #7f8c8d; }
        """)
        clear_btn.clicked.connect(self.clear_filters)

        filter_layout.addWidget(QLabel("پلاک:"))
        filter_layout.addWidget(self.search_input)
        filter_layout.addWidget(QLabel("از:"))
        filter_layout.addWidget(self.date_from)
        filter_layout.addWidget(QLabel("تا:"))
        filter_layout.addWidget(self.date_to)
        filter_layout.addWidget(QLabel("نوع:"))
        filter_layout.addWidget(self.type_filter)
        filter_layout.addWidget(search_btn)
        filter_layout.addWidget(clear_btn)

        filter_group.setLayout(filter_layout)
        layout.addWidget(filter_group)

        # جدول
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels([
            "ردیف", "پلاک", "نوع", "ورود", "خروج",
            "مدت", "هزینه", "تخفیف"
        ])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                border: 1px solid #ddd;
                border-radius: 5px;
                font-size: 13px;
                alternate-background-color: #f8f9fa;
            }
            QHeaderView::section {
                background-color: #2c3e50;
                color: white;
                padding: 10px;
                font-weight: bold;
            }
            QTableWidget::item { padding: 8px; }
        """)

        layout.addWidget(self.table)

        # صفحه‌بندی
        page_layout = QHBoxLayout()

        self.prev_btn = QPushButton("◀ قبلی")
        self.prev_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 10px 25px;
                border-radius: 6px;
                font-weight: bold;
                border: none;
            }
            QPushButton:hover { background-color: #2980b9; }
        """)
        self.prev_btn.clicked.connect(self.prev_page)

        self.page_label = QLabel("صفحه ۱")
        self.page_label.setAlignment(Qt.AlignCenter)
        self.page_label.setStyleSheet("font-size: 15px; font-weight: bold; color: #2c3e50;")

        self.next_btn = QPushButton("بعدی ▶")
        self.next_btn.setStyleSheet("""
            QPushButton {
                background-color: #3498db;
                color: white;
                padding: 10px 25px;
                border-radius: 6px;
                font-weight: bold;
                border: none;
            }
            QPushButton:hover { background-color: #2980b9; }
        """)
        self.next_btn.clicked.connect(self.next_page)

        page_layout.addWidget(self.prev_btn)
        page_layout.addWidget(self.page_label)
        page_layout.addWidget(self.next_btn)

        layout.addLayout(page_layout)

        # دکمه خروجی
        export_btn = QPushButton("📥 خروجی Excel")
        export_btn.setStyleSheet("""
            QPushButton {
                background-color: #27ae60;
                color: white;
                font-weight: bold;
                padding: 15px 30px;
                border-radius: 8px;
                font-size: 15px;
                border: none;
            }
            QPushButton:hover { background-color: #2ecc71; }
        """)
        export_btn.clicked.connect(self.export_excel)

        layout.addWidget(export_btn)

        spacer = QWidget()
        spacer.setMinimumHeight(30)
        layout.addWidget(spacer)

    def load_data(self):
        """بارگذاری داده‌ها"""
        try:
            plate_number = self.search_input.text().strip() or None
            date_from = self.date_from.date().toString("yyyy-MM-dd")
            date_to = self.date_to.date().toString("yyyy-MM-dd")

            type_map = {"همه": None, "شخصی": "personal", "تاکسی": "taxi",
                       "دولتی": "governmental", "نظامی": "military", "سیاسی": "diplomatic"}
            plate_type = type_map.get(self.type_filter.currentText())

            result = self.db.get_history(
                page=self.current_page,
                per_page=self.per_page,
                date_from=date_from,
                date_to=date_to,
                plate_number=plate_number,
                plate_type=plate_type
            )

            self.display_records(result['records'])

            total_pages = result['total_pages']
            self.page_label.setText(f"صفحه {self.current_page} از {total_pages}")
            self.prev_btn.setEnabled(self.current_page > 1)
            self.next_btn.setEnabled(self.current_page < total_pages)

        except Exception as e:
            QMessageBox.warning(self, "خطا", str(e))

    def display_records(self, records):
        """نمایش رکوردها"""
        self.table.setRowCount(0)

        for i, record in enumerate(records):
            self.table.insertRow(i)

            plate = IranianPlate.from_full_plate(record['plate_number'])

            self.table.setItem(i, 0, QTableWidgetItem(str((self.current_page - 1) * self.per_page + i + 1)))

            plate_item = QTableWidgetItem(plate.display_format)
            plate_item.setFont(QFont("Tahoma", 11, QFont.Bold))
            self.table.setItem(i, 1, plate_item)

            self.table.setItem(i, 2, QTableWidgetItem(plate.type_display))
            self.table.setItem(i, 3, QTableWidgetItem(record['entry_time'][:16]))
            self.table.setItem(i, 4, QTableWidgetItem(record['exit_time'][:16]))
            self.table.setItem(i, 5, QTableWidgetItem(f"{record['duration_hours']:.1f} ساعت"))

            cost_item = QTableWidgetItem(f"{record['final_cost']:,.0f} تومان")
            cost_item.setForeground(QColor("#e74c3c" if record['final_cost'] > 0 else "#27ae60"))
            self.table.setItem(i, 6, cost_item)

            discount_text = f"{record['discount_percent']}%" if record['discount_percent'] > 0 else "-"
            self.table.setItem(i, 7, QTableWidgetItem(discount_text))

    def prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.load_data()

    def next_page(self):
        self.current_page += 1
        self.load_data()

    def clear_filters(self):
        self.search_input.clear()
        self.date_from.setDate(QDate.currentDate().addDays(-7))
        self.date_to.setDate(QDate.currentDate())
        self.type_filter.setCurrentIndex(0)
        self.current_page = 1
        self.load_data()

    def export_excel(self):
        """خروجی Excel"""
        try:
            result = self.db.get_history(
                date_from=self.date_from.date().toString("yyyy-MM-dd"),
                date_to=self.date_to.date().toString("yyyy-MM-dd"),
                per_page=10000
            )

            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "تاریخچه پارکینگ"
            ws.sheet_view.rightToLeft = True

            headers = ['ردیف', 'پلاک', 'نوع', 'استان', 'ورود', 'خروج', 'مدت', 'هزینه']

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(name='Tahoma', size=11, bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2c3e50', end_color='2c3e50', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for i, r in enumerate(result['records'], 1):
                ws.cell(row=i+1, column=1, value=i)
                ws.cell(row=i+1, column=2, value=r['plate_number'])
                ws.cell(row=i+1, column=3, value=r.get('plate_type', 'شخصی'))
                ws.cell(row=i+1, column=4, value=r.get('province', ''))
                ws.cell(row=i+1, column=5, value=r['entry_time'][:16])
                ws.cell(row=i+1, column=6, value=r['exit_time'][:16])
                ws.cell(row=i+1, column=7, value=f"{r['duration_hours']:.1f}")
                ws.cell(row=i+1, column=8, value=f"{r['final_cost']:,.0f}")

            for col in range(1, len(headers) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

            os.makedirs('exports', exist_ok=True)
            filename = f"exports/history_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(filename)

            QMessageBox.information(self, "موفق", f"✅ فایل در {filename} ذخیره شد")

        except ImportError:
            QMessageBox.warning(self, "خطا", "کتابخانه openpyxl نصب نیست\npip install openpyxl")